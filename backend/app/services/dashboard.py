import datetime
import json
from typing import Any

from sqlalchemy import func, extract
from sqlalchemy.orm import Session

from app.models.models import DailySummary, ReconResult, ExceptionDetail, SummaryRow, ReconPair, ExpectedFile, FileReceipt, ImportBatch


EXCEPTION_LABELS = {
    "PRICE_MISMATCH": "Pricing Difference",
    "DANA_ONLY_EXT_CHECK": None,  # mapped dynamically
    "DB_ONLY_EXT_CHECK": None,    # mapped dynamically
    "FORCE_FAILED": "Force Failed",
}


def _find_metric(rows: list[SummaryRow], keywords: list[str], unit: str, source: str) -> float:
    for row in rows:
        desc = (row.description or "").upper().strip()
        if not any(kw in desc for kw in (k.upper() for k in keywords)):
            continue
        row_unit = (row.unit or "").upper().strip()
        if row_unit != unit.upper():
            continue
        if source == "bas":
            return row.bas_value or 0.0
        elif source == "dana":
            return row.dana_value or 0.0
        elif source == "diff":
            return row.chksum_value or 0.0
    return 0.0


def get_overview(db: Session, start_date=None, end_date=None, pair_id: int | None = None):
    pair = None
    if pair_id is not None:
        pair = db.query(ReconPair).filter(ReconPair.id == pair_id).first()

    date_q = db.query(SummaryRow.trx_date)
    if start_date:
        date_q = date_q.filter(SummaryRow.trx_date >= start_date)
    if end_date:
        date_q = date_q.filter(SummaryRow.trx_date <= end_date)
    if pair_id is not None:
        date_q = date_q.filter(SummaryRow.recon_pair_id == pair_id)

    latest_date = date_q.order_by(SummaryRow.trx_date.desc()).limit(1).scalar()

    source_a = pair.source_a if pair else "BAS"
    source_b = pair.source_b if pair else "DANA"
    pair_code = pair.pair_code if pair else None
    pair_name = pair.pair_name if pair else None

    labels = {
        "DANA_ONLY_EXT_CHECK": f"Missing in {source_a}",
        "DB_ONLY_EXT_CHECK": f"Missing in {source_b}",
    }

    if latest_date:
        q = db.query(SummaryRow).filter(SummaryRow.trx_date == latest_date)
        if pair_id is not None:
            q = q.filter(SummaryRow.recon_pair_id == pair_id)
        rows = q.order_by(SummaryRow.row_order).all()

        if pair_id is None and rows and not pair:
            first_pair_id = rows[0].recon_pair_id
            if first_pair_id:
                pair = db.query(ReconPair).filter(ReconPair.id == first_pair_id).first()
                if pair:
                    source_a = pair.source_a
                    source_b = pair.source_b
                    pair_code = pair.pair_code
                    pair_name = pair.pair_name
                    labels = {
                        "DANA_ONLY_EXT_CHECK": f"Missing in {source_a}",
                        "DB_ONLY_EXT_CHECK": f"Missing in {source_b}",
                    }

        total_trx_a = _find_metric(rows, ["STATUS: SUCCESS"], "#", "bas")
        total_trx_b = _find_metric(rows, ["STATUS: SUCCESS"], "#", "dana")
        total_nominal_a = _find_metric(rows, ["STATUS: SUCCESS"], "RP.", "bas")
        total_nominal_b = _find_metric(rows, ["STATUS: SUCCESS"], "RP.", "dana")
        settlement_a = _find_metric(rows, ["TOTAL SETTLEMENT"], "RP.", "bas")
        settlement_b = _find_metric(rows, ["TOTAL SETTLEMENT"], "RP.", "dana")
        diff = (
            (_find_metric(rows, ["BEDA HARGA"], "RP.", "dana") - _find_metric(rows, ["BEDA HARGA"], "RP.", "bas"))
            + _find_metric(rows, ["ADA DI DANA TIDAK ADA DI"], "RP.", "dana")
            + (-_find_metric(rows, ["TIDAK ADA DI DANA"], "RP.", "bas"))
        )
        max_settlement = max(settlement_a, settlement_b)
        diff_pct = round((abs(diff) / max_settlement * 100) if max_settlement > 0 else 0, 2)

        source_b_settlement_total: int | None = None
        source_b_file_name: str | None = None
        if pair and latest_date:
            source_b_file = (
                db.query(ExpectedFile)
                .filter(ExpectedFile.recon_pair_id == pair.id, ExpectedFile.source == source_b, ExpectedFile.active.is_(True))
                .first()
            )
            if source_b_file:
                h1_date = latest_date + datetime.timedelta(days=1)
                receipt = (
                    db.query(FileReceipt)
                    .filter(FileReceipt.expected_file_id == source_b_file.id, FileReceipt.file_date == h1_date)
                    .order_by(FileReceipt.created_at.desc())
                    .first()
                )
                if receipt and receipt.import_batch:
                    source_b_settlement_total = receipt.import_batch.source_settlement_total
                    source_b_file_name = receipt.file_name

        summary_amounts: dict[str, float] = {
            "PRICE_MISMATCH": (
                _find_metric(rows, ["BEDA HARGA"], "RP.", "dana")
                - _find_metric(rows, ["BEDA HARGA"], "RP.", "bas")
            ),
            "ONLY_IN_DANA": _find_metric(rows, ["ADA DI DANA TIDAK ADA DI"], "RP.", "dana"),
            "ONLY_IN_DB": -_find_metric(rows, ["TIDAK ADA DI DANA"], "RP.", "bas"),
            "FORCE_FAILED": -_find_metric(rows, ["FORCE FAILED"], "RP.", "bas"),
            "DANA_ONLY_EXT_CHECK": _find_metric(rows, ["ADA DI DANA TIDAK ADA DI"], "RP.", "dana"),
            "DB_ONLY_EXT_CHECK": -_find_metric(rows, ["TIDAK ADA DI DANA"], "RP.", "bas"),
        }

        ds_q = db.query(DailySummary.id).filter(DailySummary.trx_date == latest_date)
        if pair_id is not None:
            ds_q = ds_q.filter(
                (DailySummary.recon_pair_id == pair_id) | (DailySummary.recon_pair_id.is_(None))
            )
        ds_ids = [r.id for r in ds_q.all()]
        exception_summaries: list[dict[str, Any]] = []
        if ds_ids:
            agg = (
                db.query(
                    ExceptionDetail.exception_type,
                    func.count(ExceptionDetail.id).label("count"),
                )
                .filter(ExceptionDetail.daily_summary_id.in_(ds_ids))
                .group_by(ExceptionDetail.exception_type)
                .all()
            )
            for et, count in agg:
                if et in ("ONLY_IN_DANA", "ONLY_IN_DB"):
                    continue
                exception_summaries.append({
                    "exception_type": et,
                    "label": (
                        labels.get(et)
                        or EXCEPTION_LABELS.get(et)
                        or et
                    ),
                    "transaction_count": count or 0,
                    "difference_amount": int(summary_amounts.get(et, 0)),
                })

        seen_types = {es["exception_type"] for es in exception_summaries}
        for et in ("DANA_ONLY_EXT_CHECK", "DB_ONLY_EXT_CHECK", "PRICE_MISMATCH", "FORCE_FAILED"):
            if et not in seen_types:
                amt = int(summary_amounts.get(et, 0))
                if amt != 0:
                    exception_summaries.append({
                        "exception_type": et,
                        "label": labels.get(et) or EXCEPTION_LABELS.get(et) or et,
                        "transaction_count": 0,
                        "difference_amount": amt,
                    })

        return {
            "trx_date": latest_date,
            "pair_code": pair_code,
            "pair_name": pair_name,
            "source_a": source_a,
            "source_b": source_b,
            "title": f"REKONSILIASI {source_a} x {source_b}",
            "columns": ["No.", "DESKRIPSI", "Unit", f"{source_a}", f"{source_b}", "CHKSUM (E-D)"],
            "total_transaction_source_a": int(total_trx_a),
            "total_nominal_source_a": int(total_nominal_a),
            "total_transaction_source_b": int(total_trx_b),
            "total_nominal_source_b": int(total_nominal_b),
            "settlement_source_a": int(settlement_a),
            "settlement_source_b": int(settlement_b),
            "settlement_difference": int(diff),
            "settlement_difference_percent": diff_pct,
            "settlement_direction": pair.settlement_direction if pair else "RECEIVABLE",
            "source_b_settlement_total": source_b_settlement_total,
            "source_b_file_name": source_b_file_name,
            "exception_summaries": exception_summaries,
            "rows": [
                {
                    "id": row.id,
                    "row_order": row.row_order,
                    "no": row.no,
                    "description": row.description,
                    "unit": row.unit,
                    "bas_value": row.bas_value,
                    "dana_value": row.dana_value,
                    "chksum_value": row.chksum_value,
                    "is_section": row.is_section,
                }
                for row in rows
            ],
        }

    return {
        "trx_date": None,
        "pair_code": pair_code,
        "pair_name": pair_name,
        "source_a": source_a,
        "source_b": source_b,
        "title": f"REKONSILIASI {source_a} x {source_b}",
        "columns": ["No.", "DESKRIPSI", "Unit", f"{source_a}", f"{source_b}", "CHKSUM (E-D)"],
        "total_transaction_source_a": 0,
        "total_nominal_source_a": 0,
        "total_transaction_source_b": 0,
        "total_nominal_source_b": 0,
        "settlement_source_a": 0,
        "settlement_source_b": 0,
        "settlement_difference": 0,
        "settlement_difference_percent": 0,
        "settlement_direction": pair.settlement_direction if pair else "RECEIVABLE",
        "source_b_settlement_total": None,
        "source_b_file_name": None,
        "exception_summaries": [],
        "rows": [],
    }


def get_daily(db: Session, start_date=None, end_date=None):
    q = db.query(
        DailySummary.trx_date,
        func.sum(DailySummary.total_transaction).label("total_transaction"),
        func.sum(DailySummary.success_transaction).label("success_transaction"),
        func.sum(DailySummary.pending_transaction).label("pending_transaction"),
        func.sum(DailySummary.failed_transaction).label("failed_transaction"),
        func.sum(DailySummary.gross_amount).label("gross_amount"),
        func.sum(DailySummary.settlement_amount).label("settlement_amount"),
        func.sum(DailySummary.difference_amount).label("difference_amount"),
    ).group_by(DailySummary.trx_date)

    if start_date:
        q = q.filter(DailySummary.trx_date >= start_date)
    if end_date:
        q = q.filter(DailySummary.trx_date <= end_date)
    q = q.order_by(DailySummary.trx_date)

    return [dict(r._mapping) for r in q.all()]


def get_weekly(db: Session, start_date=None, end_date=None):
    q = db.query(
        func.strftime("%Y-%W", DailySummary.trx_date).label("week"),
        func.sum(DailySummary.total_transaction).label("total_transaction"),
        func.sum(DailySummary.success_transaction).label("success_transaction"),
        func.sum(DailySummary.pending_transaction).label("pending_transaction"),
        func.sum(DailySummary.failed_transaction).label("failed_transaction"),
        func.sum(DailySummary.gross_amount).label("gross_amount"),
        func.sum(DailySummary.settlement_amount).label("settlement_amount"),
        func.sum(DailySummary.difference_amount).label("difference_amount"),
    ).group_by("week")

    if start_date:
        q = q.filter(DailySummary.trx_date >= start_date)
    if end_date:
        q = q.filter(DailySummary.trx_date <= end_date)
    q = q.order_by("week")

    return [dict(r._mapping) for r in q.all()]


def get_monthly(db: Session, start_date=None, end_date=None):
    q = db.query(
        func.strftime("%Y-%m", DailySummary.trx_date).label("month"),
        func.sum(DailySummary.total_transaction).label("total_transaction"),
        func.sum(DailySummary.success_transaction).label("success_transaction"),
        func.sum(DailySummary.pending_transaction).label("pending_transaction"),
        func.sum(DailySummary.failed_transaction).label("failed_transaction"),
        func.sum(DailySummary.gross_amount).label("gross_amount"),
        func.sum(DailySummary.settlement_amount).label("settlement_amount"),
        func.sum(DailySummary.difference_amount).label("difference_amount"),
    ).group_by("month")

    if start_date:
        q = q.filter(DailySummary.trx_date >= start_date)
    if end_date:
        q = q.filter(DailySummary.trx_date <= end_date)
    q = q.order_by("month")

    return [dict(r._mapping) for r in q.all()]


def get_trend(db: Session, start_date=None, end_date=None):
    daily = get_daily(db, start_date, end_date)
    return [
        {
            "label": str(d["trx_date"]),
            "total": d["total_transaction"],
            "success": d["success_transaction"],
            "pending": d["pending_transaction"],
            "failed": d["failed_transaction"],
        }
        for d in daily
    ]


def get_recon(db: Session, start_date=None, end_date=None):
    q = db.query(ReconResult)
    if start_date:
        q = q.join(DailySummary).filter(DailySummary.trx_date >= start_date)
    if end_date:
        q = q.join(DailySummary).filter(DailySummary.trx_date <= end_date)
    return q.all()


def get_drilldown(
    db: Session,
    trx_date: datetime.date,
    exception_type: str | None = None,
    q: str | None = None,
    limit: int = 500,
    offset: int = 0,
    source_a: str | None = None,
    source_b: str | None = None,
    pair_id: int | None = None,
):
    ds_ids = [
        row.id for row in db.query(DailySummary.id).filter(DailySummary.trx_date == trx_date).all()
    ]
    if not ds_ids:
        return {"trx_date": trx_date, "exceptions": [], "total": 0, "limit": limit, "offset": offset}

    query = db.query(ExceptionDetail).filter(ExceptionDetail.daily_summary_id.in_(ds_ids))
    if exception_type:
        query = query.filter(ExceptionDetail.exception_type == exception_type)
    if q:
        like = f"%{q}%"
        query = query.filter(
            ExceptionDetail.reference_number.ilike(like)
            | ExceptionDetail.product_code.ilike(like)
            | ExceptionDetail.raw_data.ilike(like)
        )

    total = query.count()
    exceptions = (
        query.order_by(ExceptionDetail.exception_type, ExceptionDetail.reference_number)
        .offset(offset)
        .limit(limit)
        .all()
    )

    columns = _extract_raw_columns(exceptions)

    sa = source_a or "BAS"
    sb = source_b or "DANA"
    exposure_specs: dict[str, tuple[list[str], str, str]] = {
        "PRICE_MISMATCH": (["BEDA HARGA"], "RP.", "diff"),
        "ONLY_IN_DANA": ([f"TIDAK ADA DI {sa}"], "RP.", "dana"),
        "ONLY_IN_DB": ([f"TIDAK ADA DI {sb}"], "RP.", "bas"),
        "FORCE_FAILED": (["FORCE FAILED"], "RP.", "bas"),
        "DANA_ONLY_EXT_CHECK": ([f"TIDAK ADA DI {sa}"], "RP.", "dana"),
        "DB_ONLY_EXT_CHECK": ([f"TIDAK ADA DI {sb}"], "RP.", "bas"),
    }
    summary_exposure: float = 0.0
    sr_q = db.query(SummaryRow).filter(SummaryRow.trx_date == trx_date)
    if pair_id is not None:
        sr_q = sr_q.filter(SummaryRow.recon_pair_id == pair_id)
    sr_rows = list(sr_q.all())
    if exception_type and exception_type in exposure_specs:
        keywords, unit, source = exposure_specs[exception_type]
        summary_exposure = int(_find_metric(sr_rows, keywords, unit, source))
    elif not exception_type:
        sa = source_a or "BAS"
        sb = source_b or "DANA"
        price_diff = _find_metric(sr_rows, ["BEDA HARGA"], "RP.", "diff")
        dana_only = _find_metric(sr_rows, [f"TIDAK ADA DI {sa}"], "RP.", "dana")
        bas_only = _find_metric(sr_rows, [f"TIDAK ADA DI {sb}"], "RP.", "bas")
        summary_exposure = int(price_diff + dana_only - bas_only)

    available_types = [
        row[0] for row in
        db.query(ExceptionDetail.exception_type.distinct())
        .filter(ExceptionDetail.daily_summary_id.in_(ds_ids))
        .filter(~ExceptionDetail.exception_type.in_(["ONLY_IN_DANA", "ONLY_IN_DB"]))
        .all()
    ]

    pricing_breakdown: list[dict[str, Any]] = []
    if exception_type in ("PRICE_MISMATCH", None):
        count_rows: dict[str, dict[str, float]] = {}
        amount_rows: dict[str, dict[str, float]] = {}
        for row in sr_rows:
            desc = (row.description or "").strip()
            if not desc or row.is_section:
                continue
            if "STATUS" in desc.upper() or "BREAKDOWN" in desc.upper():
                continue
            sku = desc
            if (row.unit or "").upper().strip() in ("#",):
                count_rows.setdefault(sku, {})["bas"] = row.bas_value or 0
                count_rows[sku]["dana"] = row.dana_value or 0
            elif (row.unit or "").upper().strip() in ("RP.", "RP"):
                amount_rows.setdefault(sku, {})["bas"] = row.bas_value or 0
                amount_rows[sku]["dana"] = row.dana_value or 0
        for sku in count_rows:
            cnt = count_rows.get(sku, {})
            amt = amount_rows.get(sku, {})
            bas_cnt = cnt.get("bas", 0)
            dana_cnt = cnt.get("dana", 0)
            bas_amt = amt.get("bas", 0)
            dana_amt = amt.get("dana", 0)
            internal_price = bas_amt / bas_cnt if bas_cnt > 0 else 0
            dana_price = dana_amt / dana_cnt if dana_cnt > 0 else 0
            diff_per_unit = dana_price - internal_price
            total_impact = dana_amt - bas_amt
            pricing_breakdown.append({
                "partner_sku": sku,
                "internal_price": int(internal_price),
                "dana_price": int(dana_price),
                "diff_per_unit": int(diff_per_unit),
                "transaction_count": int(bas_cnt),
                "total_impact": int(total_impact),
            })
        pricing_breakdown.sort(key=lambda x: abs(x["total_impact"]), reverse=True)

    return {
        "trx_date": trx_date,
        "columns": columns,
        "summary_exposure": int(summary_exposure),
        "exceptions": [
            {
                "id": e.id,
                "exception_type": e.exception_type,
                "reference_number": e.reference_number,
                "product_code": e.product_code,
                "amount": e.amount,
                "reason": e.reason,
                    "created_at": e.created_at,
                "raw_data": json.loads(e.raw_data) if e.raw_data else None,
                "bas_value": _extract_bas(e, json.loads(e.raw_data)) if e.raw_data else 0,
                "dana_value": _extract_dana(e, json.loads(e.raw_data)) if e.raw_data else 0,
                "diff_value": _extract_diff(e, json.loads(e.raw_data)) if e.raw_data else 0,
            }
            for e in exceptions
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
        "available_types": available_types,
        "pricing_breakdown": pricing_breakdown,
    }


def export_drilldown(
    db: Session,
    trx_date: datetime.date,
    exception_type: str | None = None,
    q: str | None = None,
    pair_id: int | None = None,
    source_a: str = "BAS",
    source_b: str = "DANA",
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    sr_q = db.query(SummaryRow).filter(SummaryRow.trx_date == trx_date)
    if pair_id is not None:
        sr_q = sr_q.filter(SummaryRow.recon_pair_id == pair_id)
    sr_rows = list(sr_q.all())

    ds_ids = [
        row.id for row in db.query(DailySummary.id).filter(DailySummary.trx_date == trx_date).all()
    ]
    if not ds_ids:
        return StreamingResponse(BytesIO(b""), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    query = db.query(ExceptionDetail).filter(ExceptionDetail.daily_summary_id.in_(ds_ids))
    if exception_type:
        query = query.filter(ExceptionDetail.exception_type == exception_type)
    if q:
        like = f"%{q}%"
        query = query.filter(
            ExceptionDetail.reference_number.ilike(like)
            | ExceptionDetail.product_code.ilike(like)
            | ExceptionDetail.raw_data.ilike(like)
        )

    exceptions = query.order_by(ExceptionDetail.exception_type, ExceptionDetail.reference_number).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Drilldown"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    raw_keys = _extract_raw_columns(exceptions)

    headers = ["#", "Reference", "Kategori", f"{source_a}", f"{source_b}", "Selisih", "Reason"] + raw_keys
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, e in enumerate(exceptions, 2):
        raw = json.loads(e.raw_data) if e.raw_data else {}
        bas_val = _extract_bas(e, raw)
        dana_val = _extract_dana(e, raw)
        diff_val = _extract_diff(e, raw)
        values = [
            row_idx - 1,
            e.reference_number or "",
            e.exception_type or "",
            bas_val,
            dana_val,
            diff_val,
            e.reason or "",
        ] + [raw.get(k, "") for k in raw_keys]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            cell.border = thin_border
            cell.font = Font(size=9)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30

    pair_label = source_a.replace("+", "-").replace(" ", "-") + "-" + source_b.replace("+", "-").replace(" ", "-")
    type_label = exception_type.replace("_", "-").lower() if exception_type else "all"
    filename = f"drilldown_{pair_label}_{trx_date}_{type_label}.xlsx"
    filename = filename.replace(" ", "_").replace("/", "-")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_float(v: object) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _extract_bas(exc: ExceptionDetail, raw: dict) -> float:
    t = exc.exception_type
    if t in ("ONLY_IN_DANA", "DANA_ONLY_EXT_CHECK", "FORCE_FAILED"):
        return 0.0
    for key in ("price", "hpp_partner", "bas"):
        val = raw.get(key)
        if val:
            return _parse_float(val)
    return 0.0


def _extract_dana(exc: ExceptionDetail, raw: dict) -> float:
    t = exc.exception_type
    if t in ("ONLY_IN_DB", "DB_ONLY_EXT_CHECK", "FORCE_FAILED"):
        return 0.0
    selisih = _parse_float(raw.get("selisih"))
    bas = _extract_bas(exc, raw)
    if bas != 0:
        return bas + selisih
    for key in ("settle", "amount", "dana"):
        val = raw.get(key)
        if val:
            return _parse_float(val)
    return 0.0


def _extract_diff(exc: ExceptionDetail, raw: dict) -> float:
    for key in ("selisih", "difference", "diff"):
        val = raw.get(key)
        if val:
            return _parse_float(val)
    bas = _extract_bas(exc, raw)
    dana = _extract_dana(exc, raw)
    return dana - bas


def _extract_raw_columns(exceptions: list) -> list[str]:
    seen: set[str] = set()
    for e in exceptions:
        if e.raw_data:
            try:
                d = json.loads(e.raw_data)
                seen.update(d.keys())
            except (json.JSONDecodeError, TypeError):
                pass
    return sorted(seen)
